import streamlit as st
import pandas as pd
from datetime import date, datetime, timedelta
import calendar
import json
import io
import time

try:
    import gspread
    from oauth2client.service_account import ServiceAccountCredentials
    HAS_GSHEETS = True
except ImportError:
    HAS_GSHEETS = False

st.set_page_config(page_title="志工排班表", page_icon="🚢", layout="wide")

INTERNAL_ZONES      = ["Z1","Z2","Z3","Z4","Z5","Z6"]
DEFAULT_ZONE_NAMES = ["1F-沉浸室劇場","1F-手扶梯驗票","2F展區、特展","3F-展區","4F-展區","5F-閱讀區"]

ADMIN_PW = st.secrets.get("admin_pw", "781223")
MAX_LOGIN_ATTEMPTS = 5

WD = {0:"一",1:"二",2:"三",3:"四",4:"五",5:"六",6:"日"}
MON_EN = ["","January","February","March","April","May","June",
           "July","August","September","October","November","December"]

st.markdown("""
<style>
#MainMenu,footer,header{visibility:hidden;}
[data-testid="stToolbar"],[data-testid="stDecoration"],
[data-testid="stElementToolbar"],
section[data-testid="stSidebar"]{display:none!important;}

/* ══ 全站白底黑字（修復 Safari 文字顯示問題） ══ */
.stApp{background:#ffffff!important;color:#000000!important;}
.block-container{
    padding:10px 6px 20px 6px!important;
    max-width:500px!important;margin:0 auto!important;
    background:#ffffff!important;
}
html,body{
    background:#ffffff!important;
    color:#000000!important;
    font-family:-apple-system,"PingFang TC","Noto Sans TC",sans-serif;
}
[class*="css"]{font-family:-apple-system,"PingFang TC","Noto Sans TC",sans-serif;}

/* 強制所有常見元素文字為黑色 */
p,span,div,label,li,td,th,h1,h2,h3,h4,h5,h6,
a,caption,strong,em,b,i,small{color:#000000!important;}

/* Streamlit markdown 容器 */
[data-testid="stMarkdownContainer"],
[data-testid="stMarkdownContainer"] *{color:#000000!important;}

/* 輸入框 */
.stTextInput input,
[data-baseweb="input"] input,
[data-baseweb="textarea"] textarea{
    color:#000000!important;
    background:#ffffff!important;
    border:1px solid #999!important;
}
.stTextInput input::placeholder,
[data-baseweb="input"] input::placeholder{color:#999999!important;}

/* 下拉選單 */
[data-testid="stSelectbox"] *,
[data-baseweb="select"] *{color:#000000!important;background:#ffffff!important;}

/* 下拉選單彈出視窗 */
[data-baseweb="popover"],[data-baseweb="popover"] *,
[data-baseweb="menu"],[data-baseweb="menu"] *,
[data-baseweb="list"],[data-baseweb="list"] *{
    color:#000000!important;background:#ffffff!important;
}
[data-baseweb="menu"] li:hover,[data-baseweb="list"] li:hover{background:#f0f0f0!important;}

/* 其他元件 */
[data-testid="stAlert"] *{color:#000000!important;}
[data-testid="stCaptionContainer"] *{color:#444444!important;}
[data-testid="stInfo"] *{color:#000000!important;}
[data-testid="stNumberInput"] input{color:#000000!important;background:#ffffff!important;}
[data-testid="stDateInput"] input{color:#000000!important;background:#ffffff!important;}
[data-testid="stTextArea"] textarea{color:#000000!important;background:#ffffff!important;}
[data-testid="stDataFrame"] *{color:#000000!important;}
button{color:#000000!important;}

.header-row{display:flex;align-items:baseline;gap:10px;margin-bottom:6px;padding-left:4px;}
.header-title{font-size:24px;font-weight:700;color:#000000;margin:0;}
.header-date{font-size:16px;font-weight:500;color:#444444;}

div[data-testid="stHorizontalBlock"]:has(>div:nth-child(7)){
    display:grid!important;grid-template-columns:repeat(7,1fr)!important;
    gap:1px!important;width:100%!important;
    margin-top:0!important;margin-bottom:0!important;
    padding-top:0!important;padding-bottom:0!important;
}
div[data-testid="stHorizontalBlock"]:has(>div:nth-child(7)) button{
    width:100%!important;min-width:0!important;padding:0!important;
    aspect-ratio:1/1!important;height:auto!important;
    display:flex;align-items:center;justify-content:center;
    line-height:1!important;border-radius:4px!important;
    border:1px solid #333333!important;font-weight:600!important;font-size:14px!important;
    margin:0!important;color:#000000!important;background:#ffffff!important;
}
div[data-testid="stHorizontalBlock"]:has(>div:nth-child(7)) ~ div[data-testid="stHorizontalBlock"]:has(>div:nth-child(7)){
    margin-top:-4px!important;
}
div[data-testid="stHorizontalBlock"]:has(>div:nth-child(7)) div[data-testid="stVerticalBlock"]{gap:0!important;}
div[data-testid="stVerticalBlockBorderWrapper"]:has(div[data-testid="stHorizontalBlock"]:has(>div:nth-child(7))){
    padding-top:0!important;padding-bottom:0!important;margin-top:0!important;margin-bottom:0!important;
}

.mnav-row{margin-bottom:4px;}
.mnav-row div[data-testid="stHorizontalBlock"]{align-items:center!important;gap:0!important;}
.mnav-row button{
    background:transparent!important;border:none!important;box-shadow:none!important;
    font-size:22px!important;font-weight:700!important;color:#000000!important;
    height:36px!important;min-width:36px!important;padding:0!important;
}
.mnav-row button:disabled{color:#cccccc!important;}

.enter-btn button{
    background:#ffffff!important;color:#000000!important;
    border:1.5px solid #000000!important;height:40px!important;
    font-size:15px!important;font-weight:700!important;margin-top:4px!important;
}

.ann-box{background:#ffffff;border:2px solid #000000;border-radius:6px;margin-top:6px!important;margin-bottom:8px!important;}
.ann-title{border-bottom:1.5px solid #000000;padding:6px;font-weight:700;text-align:center;font-size:15px;color:#000000!important;}
.ann-body{padding:8px 12px;font-size:13px;line-height:1.5;color:#000000!important;white-space:pre-wrap;}

.admin-tiny button{background:transparent!important;color:#666666!important;border:none!important;
    font-size:11px!important;padding:0!important;height:auto!important;box-shadow:none!important;}

.wk-wrap{overflow-x:auto;margin:0 0 2px 0;}
.wk-tbl{border-collapse:collapse;width:100%;font-size:12px;table-layout:fixed;}
.wk-tbl th{border:1px solid #000000;padding:2px;text-align:center;background:#eeeeee;
    font-weight:600;white-space:normal!important;word-wrap:break-word!important;
    vertical-align:middle;height:35px;font-size:11px;color:#000000!important;}
.wk-tbl td{border:1px solid #000000;padding:2px;text-align:center;vertical-align:middle;height:35px;color:#000000!important;}
.wk-date-cell{background:#f5f5f5;font-weight:700;font-size:11px;width:35px;color:#000000!important;}
.wk-shift-cell{background:#e8e8e8;font-size:10px;width:20px;font-weight:600;
    writing-mode:vertical-rl;text-orientation:upright;letter-spacing:1px;padding:0 2px;color:#000000!important;}
.wk-filled-cell{background:#FFD700;color:#000000!important;}
.wk-empty-cell{background:#ffffff;color:#000000!important;}
.wk-closed-cell{background:#e0e0e0;color:#444444!important;font-size:10px;letter-spacing:1px;
    background-image:repeating-linear-gradient(45deg,transparent,transparent 5px,#ccc 5px,#ccc 6px);}
.wk-outrange-cell{background:#d0d0d0;
    background-image:repeating-linear-gradient(45deg,transparent,transparent 4px,#aaa 4px,#aaa 5px);}
.vol-name{font-size:13px;font-weight:600;color:#000000!important;display:block;
    line-height:1.1;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;}
.sel-border{outline:2px solid #cc0000;outline-offset:-2px;}

div[data-testid="stHorizontalBlock"]:has(>div:nth-child(2):last-child){
    gap:4px!important;margin-top:2px!important;margin-bottom:2px!important;
}
div[data-testid="stHorizontalBlock"]:has(>div:nth-child(2):last-child) button{
    background:#ffffff!important;color:#000000!important;
    border:1.5px solid #555555!important;border-radius:8px!important;
    height:38px!important;font-size:11px!important;font-weight:600!important;width:100%!important;
}
div[data-testid="stHorizontalBlock"]:has(>div:nth-child(2):last-child) button:disabled{
    background:#eeeeee!important;color:#aaaaaa!important;border-color:#dddddd!important;
}

div[data-testid="stSelectbox"] label,div[data-testid="stTextInput"] label{
    font-size:13px!important;margin-bottom:0!important;
    padding-bottom:0!important;line-height:1.3!important;min-height:0!important;
    color:#000000!important;
}
div[data-testid="stSelectbox"] div[data-baseweb="select"],
div[data-testid="stTextInput"] div[data-baseweb="input"]{
    min-height:32px!important;height:32px!important;font-size:14px!important;
    background:#ffffff!important;
}
div[data-testid="stSelectbox"],div[data-testid="stTextInput"]{
    margin-bottom:3px!important;margin-top:0!important;
}

button[kind="primary"]{background:#ef4444!important;color:#ffffff!important;border:none!important;}
button:disabled{background:#e5e5e5!important;color:#888888!important;opacity:0.6!important;}
.day-header{text-align:center;font-size:12px;font-weight:700;color:#333333;margin-bottom:2px;}
.day-header.sun{color:#cc0000;}

.admin-card{background:#ffffff;border-radius:14px;padding:28px 20px 20px;
    box-shadow:0 2px 14px rgba(0,0,0,.10);}
.admin-title{color:#e53e3e;text-align:center;font-size:26px;font-weight:700;margin-bottom:24px;}
.admin-big-btn button{background:#4ECDC4!important;color:#000000!important;border:none!important;
    border-radius:10px!important;height:64px!important;font-size:18px!important;font-weight:600!important;}
.admin-back-btn button{background:#c8c8c8!important;color:#333333!important;border:none!important;
    border-radius:10px!important;height:48px!important;font-size:15px!important;}

.mini-cal-wrap{background:#ffffff;border-radius:10px;padding:10px 8px 8px;
    border:1px solid #333333;margin-bottom:10px;}
.mini-cal-month{text-align:center;font-weight:700;font-size:14px;margin-bottom:6px;color:#000000!important;}
.mini-cal-tbl{width:100%;border-collapse:collapse;table-layout:fixed;}
.mini-cal-tbl th{font-size:10px;font-weight:600;color:#555555!important;text-align:center;padding:2px 0 4px;}
.mini-cal-tbl th.mc-sun{color:#cc0000!important;}
.mini-cal-tbl td{text-align:center;padding:2px 0;}
.mc-day{width:30px;height:30px;border-radius:50%;display:inline-flex;
    align-items:center;justify-content:center;font-size:12px;font-weight:500;margin:auto;}
.mc-normal{color:#000000!important;background:transparent;}
.mc-closed-def{background:#e0e0e0;color:#555555!important;}
.mc-closed-sp{background:#ef4444;color:#ffffff!important;}
.mc-open-sp{background:#4ECDC4;color:#000000!important;}
.mc-pad{color:#cccccc!important;font-size:12px;}
.cal-legend{display:flex;flex-wrap:wrap;gap:8px;font-size:11px;margin-bottom:8px;align-items:center;color:#000000!important;}
.leg-dot{width:12px;height:12px;border-radius:50%;display:inline-block;margin-right:3px;vertical-align:middle;}

.cancel-hint{background:#fff8e1;border:1.5px solid #f59e0b;border-radius:8px;
    padding:8px 12px;font-size:12px;color:#7a3800!important;margin-bottom:6px;line-height:1.5;}
</style>
""", unsafe_allow_html=True)

# ── Google Sheets ──────────────────────────────────────────
@st.cache_resource
def init_connection():
    if not HAS_GSHEETS: return None
    if "textkey" not in st.secrets: return None
    key_dict = st.secrets["textkey"]
    scope = ["https://spreadsheets.google.com/feeds","https://www.googleapis.com/auth/drive"]
    creds = ServiceAccountCredentials.from_json_keyfile_dict(key_dict, scope)
    return gspread.authorize(creds)

CHUNK_SIZE = 40000

@st.cache_data(ttl=10)
def load_data():
    try:
        client = init_connection()
        if client is None: return {}
        sheet = client.open("volunteer_db").sheet1
        df = pd.DataFrame(sheet.get_all_records())
        raw = {}
        if not df.empty:
            df.columns = [str(c).lower() for c in df.columns]
            if "key" in df.columns and "value" in df.columns:
                for _, row in df.iterrows():
                    raw[str(row["key"])] = str(row["value"])
        assembled = {}
        chunk_keys = [k for k in raw if "__chunk_" in k]
        base_keys  = set(k.split("__chunk_")[0] for k in chunk_keys)
        for base in base_keys:
            parts = []
            i = 0
            while True:
                ck = f"{base}__chunk_{i}"
                if ck in raw:
                    parts.append(raw[ck])
                    i += 1
                else:
                    break
            if parts:
                assembled[base] = "".join(parts)
        result = {k: v for k, v in raw.items() if "__chunk_" not in k}
        result.update(assembled)
        return result
    except:
        return {}

def save_data(key, value):
    try:
        client = init_connection()
        if client is None: return
        sheet = client.open("volunteer_db").sheet1
        all_vals = sheet.get_all_values()
        key_to_row = {str(r[0]): idx+1 for idx, r in enumerate(all_vals) if r}

        if len(str(value)) <= CHUNK_SIZE:
            chunk_rows = sorted(
                [row for k, row in key_to_row.items() if k.startswith(f"{key}__chunk_")],
                reverse=True)
            for r in chunk_rows:
                sheet.delete_rows(r)
                all_vals = sheet.get_all_values()
                key_to_row = {str(r2[0]): idx+1 for idx, r2 in enumerate(all_vals) if r2}
            if key in key_to_row:
                sheet.update_cell(key_to_row[key], 2, value)
            else:
                sheet.append_row([key, value])
        else:
            chunks = [value[i:i+CHUNK_SIZE] for i in range(0, len(value), CHUNK_SIZE)]
            if key in key_to_row:
                sheet.delete_rows(key_to_row[key])
                all_vals = sheet.get_all_values()
                key_to_row = {str(r[0]): idx+1 for idx, r in enumerate(all_vals) if r}
            for ci, chunk in enumerate(chunks):
                ck = f"{key}__chunk_{ci}"
                if ck in key_to_row:
                    sheet.update_cell(key_to_row[ck], 2, chunk)
                else:
                    sheet.append_row([ck, chunk])
                    all_vals = sheet.get_all_values()
                    key_to_row = {str(r[0]): idx+1 for idx, r in enumerate(all_vals) if r}
    except Exception as e:
        st.error(f"❌ 存檔失敗: {e}")

# ── State ──────────────────────────────────────────────────
def init_state():
    if "app_ready" in st.session_state: return
    try:
        raw = load_data()
        st.session_state.bookings = raw
        try: st.session_state.open_months_list = [(m[0],m[1]) for m in json.loads(raw.get("SYS_OPEN_MONTHS","[[2026,3]]"))]
        except: st.session_state.open_months_list = [(2026,3)]
        try: st.session_state.closed_days = [datetime.strptime(d,"%Y-%m-%d").date() for d in json.loads(raw.get("SYS_CLOSED_DAYS","[]"))]
        except: st.session_state.closed_days = []
        try: st.session_state.open_days = [datetime.strptime(d,"%Y-%m-%d").date() for d in json.loads(raw.get("SYS_OPEN_DAYS","[]"))]
        except: st.session_state.open_days = []
        try: st.session_state.zone_names = json.loads(raw.get("SYS_ZONE_NAMES",json.dumps(DEFAULT_ZONE_NAMES)))
        except: st.session_state.zone_names = DEFAULT_ZONE_NAMES
        try:
            raw_vol = json.loads(raw.get("SYS_VOLUNTEERS","[]"))
            vols = []
            for v in raw_vol:
                if isinstance(v, str): vols.append({"name": v, "id": ""})
                else: vols.append(v)
            st.session_state.volunteers = vols
        except:
            st.session_state.volunteers = []
        st.session_state.announcement   = raw.get("SYS_ANNOUNCEMENT","歡迎！點選週次進行排班。")
        try: st.session_state.duty_files = json.loads(raw.get("SYS_DUTY_FILES","[]"))
        except: st.session_state.duty_files = []
        for df_meta in st.session_state.duty_files:
            dk = df_meta.get("key","")
            if dk and dk in raw:
                st.session_state.bookings[dk] = raw[dk]
        st.session_state.page           = "calendar"
        st.session_state.month_idx      = 0
        st.session_state.sel_week_start = None
        st.session_state.sel_cell       = None
        st.session_state.is_admin_auth  = False
        st.session_state.login_attempts = 0
        st.session_state.app_ready      = True
    except Exception as e:
        st.error(f"❌ 程式初始化失敗，請重新整理頁面。錯誤：{e}")
        st.stop()

init_state()

def require_admin():
    if not st.session_state.get("is_admin_auth", False):
        nav("admin_login")
        st.stop()

# ── Helpers ────────────────────────────────────────────────
def is_open(d: date) -> bool:
    if d in st.session_state.closed_days: return False
    if d in st.session_state.open_days:   return True
    if d.weekday() == 0: return False
    return True

def nav(page):
    st.session_state.page = page
    st.rerun()

def week_start(d: date) -> date:
    return d - timedelta(days=d.weekday())

def get_weeks(year, month):
    first = date(year, month, 1)
    last  = date(year, month, calendar.monthrange(year, month)[1])
    ws = week_start(first)
    weeks = []
    while ws <= last:
        weeks.append((ws, [ws+timedelta(days=i) for i in range(7)]))
        ws += timedelta(weeks=1)
    return weeks

def open_bounds():
    months = sorted(st.session_state.open_months_list)
    if not months:
        t = date.today(); return t, t
    y0,m0 = months[0];  y1,m1 = months[-1]
    return date(y0,m0,1), date(y1,m1,calendar.monthrange(y1,m1)[1])

def clipped_week_label(ws: date) -> str:
    min_d, max_d = open_bounds()
    we = ws + timedelta(days=6)
    eff_s = max(ws, min_d); eff_e = min(we, max_d)
    if eff_s == eff_e: return f"{eff_s.month}/{eff_s.day}({WD[eff_s.weekday()]})"
    return f"{eff_s.month}/{eff_s.day}({WD[eff_s.weekday()]})～{eff_e.month}/{eff_e.day}({WD[eff_e.weekday()]})"

def day_status(d: date, min_d: date, max_d: date) -> str:
    if d < min_d or d > max_d: return "outrange"
    if not is_open(d): return "closed"
    return "open"

def get_volunteer_index() -> dict:
    return {v["name"]: v for v in st.session_state.get("volunteers", [])}

def find_volunteer_by_name(name: str):
    return get_volunteer_index().get(name)

def id_matches(volunteer: dict, id_input: str) -> bool:
    vid = volunteer.get("id","").strip()
    if not vid: return False
    inp = id_input.strip()
    if not inp: return False
    return (vid[0].upper() + vid[1:]) == (inp[0].upper() + inp[1:])

# ── Page: Calendar ─────────────────────────────────────────
def page_calendar():
    months = sorted(st.session_state.open_months_list)
    if months:
        idx = min(st.session_state.month_idx, len(months)-1)
        year, month = months[idx]
    else:
        idx, year, month = 0, date.today().year, date.today().month

    if not months:
        st.warning("⚠️ 暫無開放月份"); _bottom_row([]); return

    weeks = get_weeks(year, month)
    sel_start = st.session_state.sel_week_start
    min_d, max_d = open_bounds()

    st.markdown('<div class="mnav-row">', unsafe_allow_html=True)
    _nc1, _nc2, _nc3 = st.columns([1, 5, 1])
    with _nc1:
        if st.button("◀", key="prev_m", disabled=(idx==0), use_container_width=True):
            st.session_state.month_idx = idx - 1
            st.session_state.sel_week_start = None; st.rerun()
    _nc2.markdown(f"<div style='text-align:center;font-weight:700;font-size:18px;"
                  f"line-height:36px;white-space:nowrap;color:#000000;'>{MON_EN[month]} {year}</div>",
                  unsafe_allow_html=True)
    with _nc3:
        if st.button("▶", key="next_m", disabled=(idx>=len(months)-1), use_container_width=True):
            st.session_state.month_idx = idx + 1
            st.session_state.sel_week_start = None; st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)

    hdr = st.columns(7)
    for i,lbl in enumerate(["Mon","Tue","Wed","Thu","Fri","Sat","Sun"]):
        cls = "day-header sun" if i==6 else "day-header"
        hdr[i].markdown(f'<div class="{cls}">{lbl}</div>', unsafe_allow_html=True)

    for ws, days in weeks:
        is_sel = (sel_start == ws)
        btn_type = "primary" if is_sel else "secondary"
        dcols = st.columns(7)
        for i,d in enumerate(days):
            with dcols[i]:
                if d.month != month:
                    st.markdown('<div style="aspect-ratio:1/1;"></div>', unsafe_allow_html=True)
                else:
                    status = day_status(d, min_d, max_d)
                    disabled = (status != "open")
                    if st.button(str(d.day), key=f"btn_{d}", type=btn_type,
                                 disabled=disabled, use_container_width=True):
                        st.session_state.sel_week_start = ws; st.rerun()

    if sel_start:
        lbl = f"進入排班　{clipped_week_label(sel_start)}"
        st.markdown('<div class="enter-btn">', unsafe_allow_html=True)
        if st.button(lbl, key="enter_grid", use_container_width=True):
            st.session_state.page = "week_grid"
            st.session_state.sel_cell = None; st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

    ann = st.session_state.announcement.replace("<","&lt;").replace(">","&gt;")
    st.markdown(f'<div class="ann-box"><div class="ann-title">公告</div>'
                f'<div class="ann-body">{ann}</div></div>', unsafe_allow_html=True)
    _bottom_row(months)

def _bottom_row(months):
    volunteers = st.session_state.get("volunteers", [])
    has_ids    = any(v.get("id","").strip() for v in volunteers)
    show_panel = bool(volunteers and has_ids)

    if show_panel:
        if st.button("📋 確認已排班 / 累計時數資訊", key="open_dl_panel", use_container_width=True):
            st.session_state.dl_panel_open = not st.session_state.get("dl_panel_open", False)
            if not st.session_state.dl_panel_open:
                st.session_state.pop("dl_verified_name", None)

        if st.session_state.get("dl_panel_open", False):
            _schedule_info_panel(months, volunteers)

    st.markdown('<div class="admin-tiny">', unsafe_allow_html=True)
    if st.button("管理員登入", key="admin_access"):
        nav("admin_login")
    st.markdown('</div>', unsafe_allow_html=True)


def _schedule_info_panel(months, volunteers):
    verified_name = st.session_state.get("dl_verified_name", None)
    verified_id   = st.session_state.get("dl_verified_id", None)

    st.markdown(
        '<div style="background:#fffbeb;border:1.5px solid #d1a84b;border-radius:8px;'
        'padding:6px 12px;margin-top:4px;margin-bottom:0;">'
        '<span style="font-weight:700;font-size:13px;color:#7a3800;">🪪 請輸入身分證字號以確認排班資訊</span>'
        '</div>', unsafe_allow_html=True)

    id_col, btn_col = st.columns([3, 1])
    with id_col:
        id_input = st.text_input("身分證字號", key="dl_id",
                                 placeholder="身分證字號（第一碼大小寫皆可）",
                                 label_visibility="collapsed")
    with btn_col:
        verify_clicked = st.button("驗證", key="dl_verify_btn", use_container_width=True)

    if verify_clicked:
        inp = id_input.strip()
        if not inp:
            st.error("請輸入身分證字號。")
            st.session_state.pop("dl_verified_name", None)
            st.session_state.pop("dl_verified_id", None)
        else:
            id_norm = inp[0].upper() + inp[1:]
            matched = None
            vol_index = get_volunteer_index()
            for v in vol_index.values():
                vid = v.get("id","").strip()
                if vid and (vid[0].upper() + vid[1:]) == id_norm:
                    matched = v; break
            if matched:
                st.session_state.dl_verified_name = matched["name"]
                st.session_state.dl_verified_id   = matched["id"].strip()[0].upper() + matched["id"].strip()[1:]
                verified_name = matched["name"]
                verified_id   = st.session_state.dl_verified_id
            else:
                st.error("❌ 身分證字號不符，請重新確認。")
                st.session_state.pop("dl_verified_name", None)
                st.session_state.pop("dl_verified_id", None)
                verified_name = None; verified_id = None

    if not verified_name:
        return

    st.markdown(
        f'<div style="color:#16a34a;font-weight:700;font-size:13px;margin:4px 0 2px;">'
        f'✅ 驗證成功：{verified_name}</div>', unsafe_allow_html=True)

    month_opts   = [(y, m) for y, m in sorted(months)]
    month_labels = [f"{y}年{m}月" for y, m in month_opts]
    m_sel = st.selectbox("開放月份", range(len(month_opts)),
                         format_func=lambda i: month_labels[i],
                         key="dl_month", label_visibility="collapsed")
    sel_y, sel_m = month_opts[m_sel]

    zone_names = st.session_state.zone_names
    bookings   = st.session_state.bookings
    d_cur = date(sel_y, sel_m, 1)
    d_end = date(sel_y, sel_m, calendar.monthrange(sel_y, sel_m)[1])
    records = []
    while d_cur <= d_end:
        d_str = d_cur.strftime("%Y-%m-%d")
        for shift in ["上午","下午"]:
            for z_id, z_name in zip(INTERNAL_ZONES, zone_names):
                k = f"{d_str}_{shift}_{z_id}_1"
                if bookings.get(k,"").strip() == verified_name:
                    records.append((d_cur, shift, z_name))
        d_cur += timedelta(days=1)

    total_hrs = len(records) * 3

    if records:
        rows_html = ""
        for i, (d_obj, shift, zone) in enumerate(records):
            bg = "#fffbeb" if i % 2 == 0 else "#ffffff"
            date_lbl = f"{d_obj.month}/{d_obj.day}&nbsp;<span style='color:#555555;font-size:11px;'>(週{WD[d_obj.weekday()]})</span>"
            rows_html += (
                f'<div style="display:flex;align-items:center;padding:6px 10px;'
                f'background:{bg};border-bottom:1px solid #e0d8c0;gap:6px;">'
                f'<span style="flex:0 0 66px;font-weight:700;font-size:13px;color:#000000;">{date_lbl}</span>'
                f'<span style="flex:0 0 30px;background:#3b82f6;color:#ffffff;border-radius:4px;'
                f'font-size:11px;font-weight:600;text-align:center;padding:2px 3px;">{shift}</span>'
                f'<span style="flex:1;font-size:12px;color:#000000;">{zone}</span>'
                f'</div>'
            )
        total_bar = (
            f'<div style="display:flex;justify-content:flex-end;align-items:center;'
            f'padding:7px 12px;background:#fef3c7;border-top:2px solid #f59e0b;">'
            f'<span style="font-size:13px;font-weight:700;color:#7a3800;">'
            f'本月預計排班總時數&nbsp;&nbsp;'
            f'<span style="font-size:20px;color:#dc2626;">{total_hrs}</span>&nbsp;小時</span></div>'
        )
        st.markdown(
            f'<div style="border:1.5px solid #f59e0b;border-radius:8px;overflow:hidden;margin-top:4px;">'
            f'{rows_html}{total_bar}</div>', unsafe_allow_html=True)
    else:
        st.markdown(
            f'<div style="background:#f9fafb;border:1px solid #cccccc;border-radius:8px;'
            f'padding:10px;text-align:center;color:#555555;font-size:13px;margin-top:4px;">'
            f'📭 {sel_y}年{sel_m}月 尚無排班記錄</div>', unsafe_allow_html=True)

    _duty_history_section(verified_name, verified_id)


def _duty_history_section(verified_name, verified_id):
    duty_files = st.session_state.get("duty_files", [])
    if not duty_files:
        return

    open_key = "duty_hist_open"
    is_open_flag = st.session_state.get(open_key, False)
    arrow    = "▲" if is_open_flag else "▼"
    if st.button(f"　查看已排班累計執勤時數　{arrow}", key="duty_hist_toggle", use_container_width=True):
        st.session_state[open_key] = not is_open_flag
        st.rerun()

    if not st.session_state.get(open_key, False):
        return

    file_labels = [f["name"] for f in duty_files]
    f_sel = st.selectbox("選擇年度", range(len(duty_files)),
                         format_func=lambda i: file_labels[i],
                         key="duty_file_sel", label_visibility="collapsed")
    chosen = duty_files[f_sel]
    raw_json = st.session_state.bookings.get(chosen["key"], "[]")
    try:
        all_rows = json.loads(raw_json)
    except:
        all_rows = []

    def id_norm(s):
        s = str(s).strip()
        return (s[0].upper() + s[1:]) if s else ""

    my_id   = id_norm(verified_id) if verified_id else ""
    my_rows = [r for r in all_rows if my_id and id_norm(r.get("id","")) == my_id]
    total_hrs  = sum(float(r.get("hours", 0)) for r in my_rows)
    total_disp = int(total_hrs) if total_hrs == int(total_hrs) else total_hrs

    if my_rows:
        st.markdown(
            f'<div style="background:#ffffff;border:1.5px solid #bfdbfe;border-radius:8px;'
            f'padding:12px 14px;margin-top:2px;">'
            f'<div style="font-size:13px;color:#1e40af;font-weight:600;margin-bottom:6px;">'
            f'{chosen["name"]}</div>'
            f'<div style="display:flex;justify-content:space-between;align-items:center;">'
            f'<span style="font-size:13px;color:#000000;">累計值勤時數</span>'
            f'<span style="font-size:22px;font-weight:700;color:#dc2626;">{total_disp} 小時</span>'
            f'</div>'
            f'<div style="font-size:12px;color:#555555;margin-top:6px;">謝謝您的付出 ～</div>'
            f'</div>', unsafe_allow_html=True)
        if st.button("⬇️ 下載個人值勤資料", key="duty_dl_btn", use_container_width=True):
            _export_duty_excel(verified_name, chosen["name"], my_rows, total_hrs)
    else:
        st.markdown(
            f'<div style="background:#f9fafb;border:1px solid #cccccc;border-radius:8px;'
            f'padding:10px;text-align:center;color:#888888;font-size:13px;margin-top:2px;">'
            f'📭 {chosen["name"]} 查無您的值勤記錄</div>', unsafe_allow_html=True)

def _export_duty_excel(vol_name, file_label, rows, total_hrs):
    out_rows = [{"姓名": r.get("name", vol_name),
                 "服務日期": r.get("date",""),
                 "時數(hr)": r.get("hours","")}
                for r in rows]
    total = {"姓名":"合計","服務日期":"","時數(hr)":
             int(total_hrs) if total_hrs == int(total_hrs) else total_hrs}
    df_out = pd.concat([pd.DataFrame(out_rows),
                        pd.DataFrame([total])], ignore_index=True)
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df_out.to_excel(writer, index=False, sheet_name="個人值勤資料")
            ws_xl = writer.sheets["個人值勤資料"]
            for col in ws_xl.columns:
                ws_xl.column_dimensions[col[0].column_letter].width = \
                    max(len(str(c.value or "")) for c in col) + 4
            for cell in ws_xl[ws_xl.max_row]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="DBEAFE")
        buf.seek(0)
        st.download_button(
            f"⬇️ 點此下載 {vol_name} {file_label} 值勤資料.xlsx",
            data=buf,
            file_name=f"{vol_name}_{file_label}值勤.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True, type="primary")
    except ImportError:
        csv_str = df_out.to_csv(index=False, encoding="utf-8-sig")
        st.download_button(
            f"⬇️ 點此下載 {vol_name} {file_label} 值勤資料.csv",
            data=csv_str.encode("utf-8-sig"),
            file_name=f"{vol_name}_{file_label}值勤.csv",
            mime="text/csv", use_container_width=True, type="primary")


def _do_export_excel(vol_name, sel_y, sel_m, records, total_hrs):
    rows = [{"日期": f"{d.month}/{d.day}(週{WD[d.weekday()]})",
             "上/下午": shift, "區域": zone, "時數(hr)": 3}
            for d, shift, zone in records]
    df_out = pd.concat([
        pd.DataFrame(rows),
        pd.DataFrame([{"日期":"合計","上/下午":"","區域":"","時數(hr)":total_hrs}])
    ], ignore_index=True)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df_out.to_excel(writer, index=False, sheet_name="個人班表")
            ws_xl = writer.sheets["個人班表"]
            for col in ws_xl.columns:
                ws_xl.column_dimensions[col[0].column_letter].width = \
                    max(len(str(c.value or "")) for c in col) + 4
            for cell in ws_xl[ws_xl.max_row]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="FFF3CD")
        buf.seek(0)
        st.download_button("⬇️ 點此下載 Excel 檔案", data=buf,
                           file_name=f"{vol_name}_{sel_y}{sel_m:02d}班表.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True, type="primary")
    except ImportError:
        csv_str = df_out.to_csv(index=False, encoding="utf-8-sig")
        st.download_button("⬇️ 點此下載 CSV 檔案", data=csv_str.encode("utf-8-sig"),
                           file_name=f"{vol_name}_{sel_y}{sel_m:02d}班表.csv",
                           mime="text/csv", use_container_width=True, type="primary")


# ── Page: Week Grid ────────────────────────────────────────
def page_week_grid():
    ws = st.session_state.sel_week_start
    if ws is None: nav("calendar"); return

    week_days  = [ws+timedelta(days=i) for i in range(7)]
    months     = sorted(st.session_state.open_months_list)
    cy,cm      = months[min(st.session_state.month_idx,len(months)-1)]
    zone_names = st.session_state.zone_names
    sel_cell   = st.session_state.get("sel_cell")
    min_d, max_d = open_bounds()

    st.markdown(f'<div class="header-row">'
                f'<span class="header-title">志工排班表</span>'
                f'<span class="header-date">{MON_EN[cm]} {cy}</span></div>', unsafe_allow_html=True)

    html = '<div class="wk-wrap"><table class="wk-tbl"><tr>'
    html += '<th class="wk-date-cell">日期</th><th class="wk-shift-cell"></th>'
    for z in zone_names: html += f'<th>{z}</th>'
    html += '</tr>'

    for day in week_days:
        d_str  = day.strftime('%Y-%m-%d')
        status = day_status(day, min_d, max_d)
        lbl    = f"{day.month}/{day.day}<br>({WD[day.weekday()]})"
        if status == "outrange": continue
        elif status == "closed":
            html += (f'<tr><td class="wk-date-cell" style="height:26px;">{lbl}</td>'
                     f'<td class="wk-shift-cell"></td>'
                     f'<td colspan="{len(INTERNAL_ZONES)}" class="wk-closed-cell" style="height:26px;">休館</td></tr>')
        else:
            html += (f'<tr><td class="wk-date-cell" style="border-bottom:none;height:28px;vertical-align:bottom;">{lbl}</td>'
                     f'<td class="wk-shift-cell">上午</td>')
            for z_id in INTERNAL_ZONES:
                k = f"{d_str}_上午_{z_id}_1"
                v = st.session_state.bookings.get(k,"").strip()
                cls = "wk-filled-cell" if v else "wk-empty-cell"
                sc  = " sel-border" if k==sel_cell else ""
                html += f'<td class="{cls}{sc}">{"<span class=vol-name>"+v+"</span>" if v else ""}</td>'
            html += '</tr>'
            html += (f'<tr><td class="wk-date-cell" style="border-top:none;height:28px;"></td>'
                     f'<td class="wk-shift-cell">下午</td>')
            for z_id in INTERNAL_ZONES:
                k = f"{d_str}_下午_{z_id}_1"
                v = st.session_state.bookings.get(k,"").strip()
                cls = "wk-filled-cell" if v else "wk-empty-cell"
                sc  = " sel-border" if k==sel_cell else ""
                html += f'<td class="{cls}{sc}">{"<span class=vol-name>"+v+"</span>" if v else ""}</td>'
            html += '</tr>'

    html += '</table></div>'
    st.markdown(html, unsafe_allow_html=True)

    prev_ws = ws - timedelta(weeks=1); next_ws = ws + timedelta(weeks=1)
    prev_ok = (prev_ws + timedelta(days=6)) >= min_d; next_ok = next_ws <= max_d
    nv1, nv2 = st.columns(2)
    with nv1:
        lbl = f"◀ {clipped_week_label(prev_ws)}" if prev_ok else "◀ （已是最早）"
        if st.button(lbl, key="prev_w", disabled=not prev_ok, use_container_width=True):
            st.session_state.sel_week_start = prev_ws; st.session_state.sel_cell = None; st.rerun()
    with nv2:
        lbl = f"{clipped_week_label(next_ws)} ▶" if next_ok else "（已是最後） ▶"
        if st.button(lbl, key="next_w", disabled=not next_ok, use_container_width=True):
            st.session_state.sel_week_start = next_ws; st.session_state.sel_cell = None; st.rerun()

    open_days = [d for d in week_days if day_status(d, min_d, max_d) == "open"]
    if open_days:
        st.markdown("**📝 登記排班**")
        ws_key = ws.strftime('%Y%m%d')
        d_opts  = [f"{d.month}/{d.day}({WD[d.weekday()]})" for d in open_days]
        d_idx   = st.selectbox("日期", range(len(open_days)), format_func=lambda i: d_opts[i], key=f"pk_d_{ws_key}")
        sel_date = open_days[d_idx]
        shifts  = ["上午","下午"]
        s_idx   = st.selectbox("時段", range(2), format_func=lambda i: shifts[i], key=f"pk_s_{ws_key}")
        sel_sf  = shifts[s_idx]
        z_idx   = st.selectbox("區域", range(len(zone_names)), format_func=lambda i: zone_names[i], key=f"pk_z_{ws_key}")
        sel_zid = INTERNAL_ZONES[z_idx]

        key = f"{sel_date.strftime('%Y-%m-%d')}_{sel_sf}_{sel_zid}_1"
        current_occupant = st.session_state.bookings.get(key, "").strip()

        volunteers  = st.session_state.get("volunteers", [])
        vol_names   = [v["name"] for v in volunteers]
        has_vol_list = bool(vol_names)

        if current_occupant:
            st.markdown(
                f'<div class="cancel-hint">'
                f'⚠️ 此欄位目前由 <b>{current_occupant}</b> 登記。<br>'
                f'若要取消排班，請在下方輸入本人身分證字號後按「儲存」。'
                f'</div>', unsafe_allow_html=True)
            field_label = "輸入身分證字號以取消排班"
            placeholder = "輸入本人身分證字號（第一碼大小寫皆可）"
        else:
            field_label = "輸入姓名以登記排班，若要取消排班請輸入當事人身分證號"
            placeholder = "輸入姓名"

        entry_val = st.text_input(field_label, key="in_n", placeholder=placeholder)

        if st.button("儲存", key="save_entry", use_container_width=True):
            entry = entry_val.strip()

            if current_occupant:
                if not entry:
                    st.error("❌ 請輸入身分證字號以取消排班。")
                    st.stop()
                occupant_vol = find_volunteer_by_name(current_occupant)
                if occupant_vol and has_vol_list:
                    if not id_matches(occupant_vol, entry):
                        st.error("❌ 身分證字號不符，無法取消他人的排班。")
                        st.stop()
                    fresh = load_data()
                    cloud = fresh.get(key,"")
                    if cloud.strip() and cloud != current_occupant:
                        st.error(f"⚠️ 資料已變動（目前為「{cloud}」），請重新操作。")
                        st.session_state.bookings[key] = cloud
                    else:
                        st.session_state.bookings[key] = ""
                        save_data(key, "")
                        st.session_state.sel_cell = key
                        st.success(f"✅ 已取消 {current_occupant} 的排班。")
                    st.rerun()
                else:
                    st.error("❌ 此志工無身分證記錄，請洽管理員處理。")
                    st.stop()
            else:
                if not entry:
                    st.error("❌ 請輸入姓名以登記排班，若要取消排班請輸入當事人身分證號。")
                    st.stop()
                if has_vol_list and entry not in vol_names:
                    st.error(f"❌ 「{entry}」不在志工名單中，請確認姓名是否正確。")
                    st.stop()
                fresh = load_data()
                cloud = fresh.get(key,"")
                old   = st.session_state.bookings.get(key,"")
                if cloud.strip() and cloud != old:
                    st.error(f"⚠️ 此格已被「{cloud}」搶先登記！")
                    st.session_state.bookings[key] = cloud
                else:
                    st.session_state.bookings[key] = entry
                    save_data(key, entry)
                    st.session_state.sel_cell = key
                    st.success("✅ 已儲存！")
                st.rerun()
    else:
        st.info("本週全部休館或不在開放月份範圍內")

    if st.button("返回月曆", key="exit_g", use_container_width=True):
        st.session_state.page = "calendar"
        st.session_state.sel_cell = None; st.rerun()

# ── Admin pages ────────────────────────────────────────────
def page_admin_login():
    st.markdown("<h2 style='color:#000000;'>管理員登入</h2>", unsafe_allow_html=True)

    attempts = st.session_state.get("login_attempts", 0)
    if attempts >= MAX_LOGIN_ATTEMPTS:
        st.error(f"❌ 嘗試次數過多（{MAX_LOGIN_ATTEMPTS} 次），請重新整理頁面後再試。")
        if st.button("返回", key="cancel_login_locked", use_container_width=True):
            st.session_state.login_attempts = 0
            nav("calendar")
        return

    pwd = st.text_input("密碼", type="password", key="pwd_in", placeholder="請輸入管理員密碼")
    c1, c2 = st.columns(2)
    with c1:
        if st.button("登入", key="do_login", type="primary", use_container_width=True):
            if pwd == ADMIN_PW:
                st.session_state.is_admin_auth = True
                st.session_state.login_attempts = 0
                nav("admin")
            else:
                st.session_state.login_attempts = attempts + 1
                remaining = MAX_LOGIN_ATTEMPTS - st.session_state.login_attempts
                st.error(f"密碼錯誤，剩餘嘗試次數：{remaining}")
    with c2:
        if st.button("返回", key="cancel_login", use_container_width=True):
            nav("calendar")

def page_admin():
    require_admin()
    st.markdown('<div class="admin-card"><div class="admin-title">管理員後台</div>', unsafe_allow_html=True)
    for label, dest in [("管理開放月份","admin_months"),("休館設定","admin_holidays"),
                        ("公告修改","admin_ann"),("區域名稱設定","admin_zones"),
                        ("👥 志工名單管理","admin_volunteers"),
                        ("📊 值勤時數檔案管理","admin_duty_files"),
                        ("📥 下載值班表 Excel","admin_export")]:
        st.markdown('<div class="admin-big-btn">', unsafe_allow_html=True)
        if st.button(label, key=f"ab_{dest}", use_container_width=True): nav(dest)
        st.markdown('</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)
    st.write("")
    c1, c2 = st.columns(2)
    with c1:
        st.markdown('<div class="admin-back-btn">', unsafe_allow_html=True)
        if st.button("🔓 登出", key="admin_logout", use_container_width=True):
            st.session_state.is_admin_auth = False
            nav("calendar")
        st.markdown('</div>', unsafe_allow_html=True)
    with c2:
        st.markdown('<div class="admin-back-btn">', unsafe_allow_html=True)
        if st.button("返回首頁", key="admin_back", use_container_width=True):
            nav("calendar")
        st.markdown('</div>', unsafe_allow_html=True)

def page_admin_months():
    require_admin()
    st.markdown("## 管理開放月份")
    cur = sorted(st.session_state.open_months_list)
    if cur: st.info("目前開放：" + "、".join([f"{y}年{m}月" for y,m in cur]))
    else:   st.warning("目前無開放月份")
    c1,c2,c3 = st.columns(3)
    ay = c1.number_input("年",2025,2030,2026,key="am_y")
    am = c2.selectbox("月",range(1,13),2,key="am_m")
    if c3.button("新增",key="add_m"):
        t=(ay,am)
        if t not in st.session_state.open_months_list:
            st.session_state.open_months_list.append(t)
            save_data("SYS_OPEN_MONTHS",json.dumps(st.session_state.open_months_list))
            st.success("✅ 已新增"); st.rerun()
    rm = st.multiselect("刪除月份",[f"{y}年{m}月" for y,m in cur])
    if st.button("🗑️ 刪除",key="rm_m"):
        for s in rm:
            y2,m2=s.replace("月","").split("年"); t=(int(y2),int(m2))
            if t in st.session_state.open_months_list: st.session_state.open_months_list.remove(t)
        save_data("SYS_OPEN_MONTHS",json.dumps(st.session_state.open_months_list)); st.rerun()
    if st.button("← 返回",key="bk_m"): nav("admin")

def render_mini_cal(year, month):
    weeks = get_weeks(year, month)
    html  = f'<div class="mini-cal-wrap"><div class="mini-cal-month">{year}年 {month}月</div>'
    html += '<table class="mini-cal-tbl"><tr>'
    for h in ["一","二","三","四","五","六"]: html += f'<th>{h}</th>'
    html += '<th class="mc-sun">日</th></tr>'
    for _, days in weeks:
        html += '<tr>'
        for d in days:
            if d.month != month:
                html += '<td><span class="mc-pad">·</span></td>'; continue
            if d in st.session_state.closed_days: cls = "mc-day mc-closed-sp"
            elif d in st.session_state.open_days: cls = "mc-day mc-open-sp"
            elif d.weekday() in (0, 6): cls = "mc-day mc-closed-def"
            else: cls = "mc-day mc-normal"
            html += f'<td><span class="{cls}">{d.day}</span></td>'
        html += '</tr>'
    html += '</table></div>'
    return html

def page_admin_holidays():
    require_admin()
    st.markdown("## 休館設定")
    st.caption("預設週一及週日休館，可額外設定特別休館/開館日。")
    st.markdown('<div class="cal-legend"><span><span class="leg-dot" style="background:#e0e0e0;"></span>預設休館</span>'
                '<span><span class="leg-dot" style="background:#ef4444;"></span>特別休館</span>'
                '<span><span class="leg-dot" style="background:#4ECDC4;"></span>特別開館</span>'
                '<span><span class="leg-dot" style="background:#fff;border:1px solid #333;"></span>正常開館</span></div>', unsafe_allow_html=True)
    for y, m in sorted(st.session_state.open_months_list):
        st.markdown(render_mini_cal(y, m), unsafe_allow_html=True)
    st.markdown("---")
    di = st.date_input("選擇日期", min_value=date(2025,1,1), key="hol_d")
    h1, h2 = st.columns(2)
    if h1.button("❌ 設為休館", key="set_cl", type="primary"):
        if di in st.session_state.open_days: st.session_state.open_days.remove(di)
        if di not in st.session_state.closed_days: st.session_state.closed_days.append(di)
        save_data("SYS_CLOSED_DAYS", json.dumps([d.strftime("%Y-%m-%d") for d in st.session_state.closed_days]))
        save_data("SYS_OPEN_DAYS",   json.dumps([d.strftime("%Y-%m-%d") for d in st.session_state.open_days]))
        st.success("✅ 已設為休館"); st.rerun()
    if h2.button("🟢 設為開館", key="set_op"):
        if di in st.session_state.closed_days: st.session_state.closed_days.remove(di)
        if di not in st.session_state.open_days: st.session_state.open_days.append(di)
        save_data("SYS_CLOSED_DAYS", json.dumps([d.strftime("%Y-%m-%d") for d in st.session_state.closed_days]))
        save_data("SYS_OPEN_DAYS",   json.dumps([d.strftime("%Y-%m-%d") for d in st.session_state.open_days]))
        st.success("✅ 已設為開館"); st.rerun()
    if st.session_state.closed_days:
        st.markdown("**特別休館日：** " + "、".join([f"{d}(週{WD[d.weekday()]})" for d in sorted(st.session_state.closed_days)]))
    if st.session_state.open_days:
        st.markdown("**特別開館日：** " + "、".join([f"{d}(週{WD[d.weekday()]})" for d in sorted(st.session_state.open_days)]))
    if st.button("← 返回", key="bk_h"): nav("admin")

def page_admin_export():
    require_admin()
    st.markdown("## 📥 下載值班表 Excel")
    zone_names = st.session_state.zone_names
    rows = []
    bookings = st.session_state.bookings
    min_d, max_d = open_bounds()
    for key, val in bookings.items():
        if key.startswith("SYS_") or not val.strip(): continue
        parts = key.split("_")
        if len(parts) != 4: continue
        d_str, shift, z_id, slot = parts
        if z_id not in INTERNAL_ZONES: continue
        try: d_obj = datetime.strptime(d_str, "%Y-%m-%d").date()
        except: continue
        if d_obj < min_d or d_obj > max_d: continue
        z_idx  = INTERNAL_ZONES.index(z_id)
        z_name = zone_names[z_idx] if z_idx < len(zone_names) else z_id
        rows.append({"日期":d_obj,"日期顯示":f"{d_obj.month}/{d_obj.day}(週{WD[d_obj.weekday()]})",
                     "星期":f"週{WD[d_obj.weekday()]}","時段":shift,"區域":z_name,"姓名":val.strip()})
    if not rows: st.info("目前開放月份內尚無任何值班登記。")
    else:
        df = pd.DataFrame(rows).sort_values(["日期","時段","區域"]).reset_index(drop=True)
        df_out = df[["日期顯示","星期","時段","區域","姓名"]].rename(columns={"日期顯示":"日期","時段":"上下午"})
        st.dataframe(df_out, use_container_width=True, hide_index=True)
        try:
            import openpyxl
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                df_out.to_excel(writer, index=False, sheet_name="值班表")
            buf.seek(0)
            st.download_button("⬇️ 下載 Excel (.xlsx)", data=buf, file_name=f"志工值班表_{date.today()}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
        except:
            csv_str = df_out.to_csv(index=False, encoding="utf-8-sig")
            st.download_button("⬇️ 下載 CSV", data=csv_str.encode("utf-8-sig"),
                               file_name=f"志工值班表_{date.today()}.csv", mime="text/csv", use_container_width=True)
    if st.button("← 返回", key="bk_ex"): nav("admin")

def page_admin_ann():
    require_admin()
    st.markdown("## 公告修改")
    ann = st.text_area("公告內容",st.session_state.announcement,height=160,key="ann_ta")
    if st.button("✅ 更新公告",key="upd_ann",type="primary"):
        st.session_state.announcement=ann
        save_data("SYS_ANNOUNCEMENT",ann)
        st.success("已更新！"); st.rerun()
    if st.button("← 返回",key="bk_ann"): nav("admin")

def page_admin_zones():
    require_admin()
    st.markdown("## 區域名稱設定")
    new_names = []
    for i in range(6): new_names.append(st.text_input(f"區域 {i+1}", value=st.session_state.zone_names[i], key=f"zn_{i}"))
    if st.button("✅ 儲存",type="primary"):
        st.session_state.zone_names = new_names
        save_data("SYS_ZONE_NAMES",json.dumps(new_names))
        st.success("已更新！"); st.rerun()
    if st.button("← 返回",key="bk_zn"): nav("admin")

def page_admin_volunteers():
    require_admin()
    st.markdown("## 👥 志工名單管理")
    st.caption("登錄姓名與身分證字號。")
    volunteers = st.session_state.get("volunteers", [])
    if volunteers:
        st.markdown(f"**目前登錄志工：共 {len(volunteers)} 人**")
        h1, h2, h3 = st.columns([3, 4, 1]); h1.markdown("**姓名**"); h2.markdown("**身分證**"); h3.markdown("**刪除**")
        st.markdown('<hr style="margin:2px 0 6px;">', unsafe_allow_html=True)
        to_delete = []
        for i, v in enumerate(volunteers):
            c1, c2, c3 = st.columns([3, 4, 1])
            c1.markdown(f'<div style="padding:4px 0;font-weight:600;color:#000000;">{v["name"]}</div>', unsafe_allow_html=True)
            vid = v.get("id","")
            masked = (vid[:3] + "***" + vid[-1]) if len(vid) >= 4 else ("***" if vid else "（未設定）")
            c2.markdown(f'<div style="padding:4px 0;color:#555555;font-size:13px;">{masked}</div>', unsafe_allow_html=True)
            if c3.button("✕", key=f"del_v_{i}"): to_delete.append(i)
        if to_delete:
            volunteers = [v for i,v in enumerate(volunteers) if i not in to_delete]
            st.session_state.volunteers = volunteers
            save_data("SYS_VOLUNTEERS", json.dumps(volunteers)); st.rerun()
    else: st.info("⚠️ 目前名單為空")
    st.markdown("---")
    a1, a2 = st.columns(2)
    new_name = a1.text_input("姓名", key="vol_new_name")
    new_id   = a2.text_input("身分證字號", key="vol_new_id", type="password")
    if st.button("＋ 新增", key="vol_add_one", use_container_width=True):
        nm, nid = new_name.strip(), new_id.strip().upper()
        if nm and not any(v["name"]==nm for v in volunteers):
            volunteers.append({"name": nm, "id": nid})
            st.session_state.volunteers = volunteers
            save_data("SYS_VOLUNTEERS", json.dumps(volunteers))
            st.success(f"✅ 已新增：{nm}"); st.rerun()
    with st.expander("📋 批次匯入（姓名,身分證 每行一筆）"):
        bulk = st.text_area("", placeholder="王小明,A123456789\n李美花,B234567890\n張雅婷", height=130, key="vol_bulk")
        if st.button("匯入", key="vol_bulk_add", use_container_width=True, type="primary"):
            added, updated = [], []
            for l in bulk.splitlines():
                p = [x.strip() for x in l.split(",",1)]; nm = p[0]
                if not nm: continue
                nid = p[1].upper() if len(p)>1 and p[1] else ""
                existing = next((v for v in volunteers if v["name"]==nm), None)
                if existing is None:
                    volunteers.append({"name":nm,"id":nid}); added.append(nm)
                elif nid and existing.get("id","") != nid:
                    existing["id"] = nid; updated.append(nm)
            if added or updated:
                st.session_state.volunteers = volunteers
                save_data("SYS_VOLUNTEERS", json.dumps(volunteers))
                msgs = []
                if added: msgs.append(f"新增 {len(added)} 位")
                if updated: msgs.append(f"更新身分證 {len(updated)} 位")
                st.success("✅ " + "、".join(msgs)); st.rerun()
            else:
                st.info("無新增或更新。")
    if volunteers and st.button("🚨 清空全部", key="vol_clear"):
        st.session_state.volunteers=[]; save_data("SYS_VOLUNTEERS","[]"); st.rerun()
    if st.button("← 返回", key="bk_vol"): nav("admin")

def page_admin_duty_files():
    require_admin()
    st.markdown("## 📊 值勤時數檔案管理")
    st.caption("上傳各年度志工值勤紀錄 Excel，志工輸入身分證後系統自動過濾顯示個人資料。")

    duty_files = st.session_state.get("duty_files", [])

    if duty_files:
        st.markdown(f"**已上傳檔案（共 {len(duty_files)} 份）**")
        to_delete = []
        for i, f in enumerate(duty_files):
            c1, c2, c3 = st.columns([4, 2, 1])
            new_label = c1.text_input("顯示名稱", value=f["name"], key=f"df_name_{i}",
                                       label_visibility="collapsed")
            row_count = len(json.loads(st.session_state.bookings.get(f["key"],"[]")))
            c2.markdown(f'<div style="padding:6px 0;font-size:12px;color:#555555;">{row_count} 筆記錄</div>',
                        unsafe_allow_html=True)
            if c3.button("✕", key=f"df_del_{i}"):
                to_delete.append(i)
            elif new_label != f["name"]:
                duty_files[i]["name"] = new_label
                st.session_state.duty_files = duty_files
                save_data("SYS_DUTY_FILES", json.dumps(duty_files))
                st.rerun()

        if to_delete:
            for i in sorted(to_delete, reverse=True):
                del_key = duty_files[i]["key"]
                save_data(del_key, "[]")
                duty_files.pop(i)
            st.session_state.duty_files = duty_files
            save_data("SYS_DUTY_FILES", json.dumps(duty_files))
            st.rerun()

        st.markdown("---")

    st.markdown("**上傳新的值勤紀錄 Excel**")
    st.caption("Excel 欄位需包含：姓名、身分證字號、服務日期起（或日期）、時數（或服務時數）")

    display_name = st.text_input("顯示名稱（例：114年服務時數）", key="df_new_name",
                                  placeholder="114年服務時數")
    uploaded = st.file_uploader("選擇 Excel 檔案 (.xlsx / .xls)", type=["xlsx","xls"],
                                 key="df_uploader")

    if uploaded and display_name.strip():
        if st.button("📤 解析並儲存", key="df_parse_btn", type="primary", use_container_width=True):
            try:
                fname  = uploaded.name.lower()
                engine = "xlrd" if fname.endswith(".xls") else "openpyxl"
                df_raw = pd.read_excel(uploaded, engine=engine)
                col_map = {}
                for col in df_raw.columns:
                    c = str(col).replace(" ","").replace("　","").replace("-","").replace("－","")
                    if any(k in c for k in ["姓名","名字"]):                              col_map.setdefault("name", col)
                    if any(k in c for k in ["身分證","身份證","證號","証號","ID","id"]):    col_map.setdefault("id", col)
                    if any(k in c for k in ["日期起","服務日期","開始日","日期"]):          col_map.setdefault("date", col)
                    if any(k in c for k in ["服務時數小時","時數小時","小時數"]):           col_map["hours_h"] = col
                    if any(k in c for k in ["服務時數分鐘","時數分鐘","分鐘數"]):           col_map["hours_m"] = col
                    if "hours_h" not in col_map and "hours_m" not in col_map:
                        if any(k in c for k in ["時數","服務時數","小時","時間"]):          col_map.setdefault("hours", col)

                has_split = "hours_h" in col_map or "hours_m" in col_map
                if has_split:
                    col_map.pop("hours", None)

                required = ["name","id","date"]
                required += [] if has_split else ["hours"]
                missing = [k for k in required if k not in col_map]
                if missing:
                    st.error(f"❌ 找不到必要欄位，請確認 Excel 包含：姓名、身分證字號、日期、時數。\n"
                             f"偵測到的欄位：{list(df_raw.columns)}")
                else:
                    records = []
                    for _, row in df_raw.iterrows():
                        nm  = str(row[col_map["name"]]).strip()
                        rid = str(row[col_map["id"]]).strip()
                        dt  = str(row[col_map["date"]]).strip()
                        if not nm or not rid or rid == "nan":
                            continue
                        if has_split:
                            try: h = float(row[col_map["hours_h"]]) if "hours_h" in col_map else 0
                            except: h = 0
                            try: m = float(row[col_map["hours_m"]]) if "hours_m" in col_map else 0
                            except: m = 0
                            hrs = h + m / 60.0
                        else:
                            try: hrs = float(row[col_map["hours"]])
                            except: hrs = 0
                        hrs = round(hrs, 1)
                        try:
                            dt = pd.to_datetime(dt).strftime("%Y/%m/%d")
                        except: pass
                        records.append({"name": nm, "id": rid, "date": dt, "hours": hrs})

                    data_key = f"SYS_DUTY_DATA_{int(time.time())}"
                    save_data(data_key, json.dumps(records, ensure_ascii=False))
                    duty_files.append({"name": display_name.strip(), "key": data_key})
                    st.session_state.duty_files = duty_files
                    st.session_state.bookings[data_key] = json.dumps(records, ensure_ascii=False)
                    save_data("SYS_DUTY_FILES", json.dumps(duty_files))
                    st.success(f"✅ 已儲存「{display_name.strip()}」，共 {len(records)} 筆記錄。")
                    st.rerun()
            except Exception as e:
                st.error(f"❌ 解析失敗：{e}")
    elif uploaded and not display_name.strip():
        st.warning("請先填寫顯示名稱。")

    st.markdown("---")
    if st.button("← 返回", key="bk_df"): nav("admin")


# ── Router ─────────────────────────────────────────────────
{
    "calendar":          page_calendar,
    "week_grid":         page_week_grid,
    "admin_login":       page_admin_login,
    "admin":             page_admin,
    "admin_months":      page_admin_months,
    "admin_holidays":    page_admin_holidays,
    "admin_ann":         page_admin_ann,
    "admin_zones":       page_admin_zones,
    "admin_volunteers":  page_admin_volunteers,
    "admin_duty_files":  page_admin_duty_files,
    "admin_export":      page_admin_export,
}.get(st.session_state.get("page","calendar"), page_calendar)()
